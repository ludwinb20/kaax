from django.shortcuts import render
import csv
from django.http import JsonResponse, HttpResponse,FileResponse
from django.views.decorators.csrf import csrf_exempt
import uuid
import os
from .serializers import VerificacionesSerializer
from .models import Transacciones_Prueba, Empresa, Verificaciones
import json
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import pandas as pd
from joblib import dump, load
from datetime import datetime,timedelta
from kaax_app.soporte import reporte, revisarplan
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

@csrf_exempt
def entrenamiento_csv(request):
    filepath = ''
    if request.method == 'POST':
        # verificar token de la empresa
        # # empresa_id = request.data['empresa_id']
        # # token = request.data['token']
        # # try:
        # #     empresa = Empresa.objects.get(id=empresa_id, token=token)
        # # except Empresa.DoesNotExist:
        # #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # plan_activo = revisarplan(empresa_id)

        # # Comprobar si el plan ha expirado
        # if not plan_activo:
        #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
        #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
        archivo = request.FILES.get('archivo')
        if archivo:
            # Procesar el archivo CSV
            reader = csv.reader(archivo.read().decode('utf-8').splitlines())
            # Realizar cualquier operación con los datos del archivo
            archivo_original = archivo.name
            extension = archivo_original.split('.')[-1]  # Obtiene la extensión del archivo'
            if extension != 'csv':
                data = {'resultado': 'Petición Incorrecta: El formato del archivo no es .csv'}
                return JsonResponse(data, status=404)
            nombre_archivo = str(uuid.uuid4()) + '.' + extension

            filepath = os.path.join('kaax/datos', nombre_archivo)
            with open(filepath, 'wb+') as destination:
                    for chunk in archivo.chunks():
                        destination.write(chunk)
        else:
            data = {'resultado': 'Petición Incorrecta: No se adjunto archivo'}
            return JsonResponse(data, status=402)

    else:
        data = {'resultado': 'Petición Incorrecta: Se esperaba metodo POST'}
        return JsonResponse(data, status=401)

    with open(filepath, 'r') as archivo_csv:
        csv_reader = csv.reader(archivo_csv)
        columnas = next(csv_reader)
        if 'ip_address' in columnas and 'email_address' in columnas and 'billing_state' in columnas and 'user_agent' in columnas and 'billing_postal' in columnas and 'phone_number' in columnas and 'EVENT_TIMESTAMP' in columnas and 'billing_address' in columnas and 'EVENT_LABEL' in columnas:
            for row in csv_reader:
                transaccion = Transacciones_Prueba(
                    ip_address=row[0],
                    email_address=row[1],
                    billing_state=row[2],
                    user_agent=row[3],
                    billing_postal=row[4],
                    phone_number=row[5],
                    EVENT_TIMESTAMP=row[6],
                    billing_address=row[7],
                    EVENT_LABEL=row[8],
                    empresa_id=1,
                    archivo=nombre_archivo,
                )
                transaccion.save()
        else: 
            data = {'resultado': 'Peticion incorrecta: El archivo .csv no contiene las columnas necesarias'}
            return JsonResponse(data, status=403)

    data = {'resultado': 'exitoso'}
    return JsonResponse(data, status=200)



@csrf_exempt
def entrenamiento_json(request):
    if request.method == 'POST':
        # verificar token de la empresa
        # # empresa_id = request.data['empresa_id']
        # # token = request.data['token']
        # # try:
        # #     empresa = Empresa.objects.get(id=empresa_id, token=token)
        # # except Empresa.DoesNotExist:
        # #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # plan_activo = revisarplan(empresa_id)

        # # Comprobar si el plan ha expirado
        # if not plan_activo:
        #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
        #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
        data = json.loads(request.body.decode('utf-8'))
        if data:
            if not all(key in data[0] for key in ('ip_address', 'email_address', 'billing_state', 'user_agent', 'billing_postal', 'phone_number', 'EVENT_TIMESTAMP', 'billing_address', 'EVENT_LABEL')):
                data = {'resultado': 'Petición Incorrecta: El objeto JSON no contiene las llaves necesarias'}
                return JsonResponse(data, status=404)

            for row in data:
                transaccion = Transacciones_Prueba(
                    ip_address=row['ip_address'],
                    email_address=row['email_address'],
                    billing_state=row['billing_state'],
                    user_agent=row['user_agent'],
                    billing_postal=row['billing_postal'],
                    phone_number=row['phone_number'],
                    EVENT_TIMESTAMP=row['EVENT_TIMESTAMP'],
                    billing_address=row['billing_address'],
                    EVENT_LABEL=row['EVENT_LABEL'],
                    empresa_id=1,
                    archivo='json'
                )
                transaccion.save()
        else:
            data = {'resultado': 'Petición Incorrecta: No se adjunto objeto JSON'}
            return JsonResponse(data, status=402)

    else:
        data = {'resultado': 'Petición Incorrecta: Se esperaba metodo POST'}
        return JsonResponse(data, status=401)

    data = {'resultado': 'exitoso'}
    return JsonResponse(data, status=200)


@csrf_exempt
@api_view(['POST'])
def verificar_transaccion(request):
    # validar campos necesarios
    required_fields_transaccion = ['ip_address', 'email_address', 'billing_state', 'user_agent', 'billing_postal', 'phone_number', 'EVENT_TIMESTAMP', 'billing_address']
    required_fields_empresa = ['empresa_id', 'token']
    
    for field in required_fields_empresa:
        if field not in request.data:
            return JsonResponse({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        
        
    for field in required_fields_transaccion:
        if field not in request.data['transaccion']:
            return JsonResponse({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        
        
    # verificar token de la empresa
    # # empresa_id = request.data['empresa_id']
    # # token = request.data['token']
    # # try:
    # #     empresa = Empresa.objects.get(id=empresa_id, token=token)
    # # except Empresa.DoesNotExist:
    # #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # plan_activo = revisarplan(empresa_id)

    # # Comprobar si el plan ha expirado
    # if not plan_activo:
    #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
    #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
    
    transaccion = request.data['transaccion']
    # cargar modelo y preprocesador
    pipe = load('kaax/pesos/decision_tree.joblib')
    
    df = pd.DataFrame(transaccion, index=[0])

    result = pipe.predict(df)



    verificacion = Verificaciones(
        ip_address=transaccion['ip_address'],
        email_address=transaccion['email_address'],
        billing_state=transaccion['billing_state'],
        user_agent=transaccion['user_agent'],
        billing_postal=transaccion['billing_postal'],
        phone_number=transaccion['phone_number'],
        EVENT_TIMESTAMP=transaccion['EVENT_TIMESTAMP'],
        billing_address=transaccion['billing_address'],
        resultado=result,
        empresa_id=request.data['empresa_id']
    )
    verificacion.save()

    # retornar resultado
    response_data = {'resultado': result}
    return Response(response_data)


@csrf_exempt
@api_view(['POST'])
def verificar_multiple(request):
    required_fields_transaccion = ['ip_address', 'email_address', 'billing_state', 'user_agent', 'billing_postal', 'phone_number', 'EVENT_TIMESTAMP', 'billing_address']
    required_fields_empresa = ['empresa_id', 'token']
    
    for field in required_fields_empresa:
        if field not in request.data:
            return JsonResponse({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

    # verificar token de la empresa
    empresa_id = request.data['empresa_id']
    token = request.data['token']
    try:
        empresa = Empresa.objects.get(id=empresa_id, token=token)
    except Empresa.DoesNotExist:
        return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
        
    transacciones = request.data['transacciones']
    print(transacciones)
    pipe = load('kaax/pesos/decision_tree.joblib')
    i = 1
    respuesta = []
    
    for transaccion in transacciones:
        for field in required_fields_transaccion:
            if field not in transaccion:
                return JsonResponse({'error': f'Falta el campo {field} en la transaccion #{i}'}, status=status.HTTP_400_BAD_REQUEST)
        df = pd.DataFrame(transaccion, index=[0])
            
        result = pipe.predict(df)

        verificacion = Verificaciones(
            ip_address=transaccion['ip_address'],
            email_address=transaccion['email_address'],
            billing_state=transaccion['billing_state'],
            user_agent=transaccion['user_agent'],
            billing_postal=transaccion['billing_postal'],
            phone_number=transaccion['phone_number'],
            EVENT_TIMESTAMP=transaccion['EVENT_TIMESTAMP'],
            billing_address=transaccion['billing_address'],
            resultado=result,
            empresa_id=request.data['empresa_id']
        )
        
        verificacion.save()
            
        arreglo = {"transaccion": i, "resultado": result[0]}
            
        respuesta.append(arreglo)
        i=i+1
        
    return Response(respuesta)

@csrf_exempt
@api_view(['GET'])
def reporte_excel(request):
    required_fields = ['fechaInicial', 'fechaFinal', 'token', 'id']
    
    for field in required_fields:
        if field not in request.data:
            return JsonResponse({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        
    
        # verificar token de la empresa
        # # empresa_id = request.data['empresa_id']
        # # token = request.data['token']
        # # try:
        # #     empresa = Empresa.objects.get(id=empresa_id, token=token)
        # # except Empresa.DoesNotExist:
        # #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # plan_activo = revisarplan(empresa_id)

        # # Comprobar si el plan ha expirado
        # if not plan_activo:
        #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
        #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
    formato_esperado = '%Y-%m-%d'
    
    try:
        datetime.strptime(request.data['fechaInicial'], formato_esperado)
        datetime.strptime(request.data['fechaFinal'], formato_esperado)
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha no valido(El formato debe ser yyyy-mm-dd, ejemplo: 2023-05-11)'}, status=status.HTTP_401_UNAUTHORIZED)
    
    datos = reporte(1, request.data[ 'fechaInicial'], request.data['fechaFinal'])
    
        # Extraer los datos relevantes
    total_registros = datos["total_registros"]
    total_fraud = datos["total_fraud"]
    porcentaje_fraud = datos["porcentaje_fraud"]
    total_legit = datos["total_legit"]
    porcentaje_legit = datos["porcentaje_legit"]
    fraud_por_hora = datos["fraud_por_hora"]
    legit_por_hora = datos["legit_por_hora"]
    fraud_por_dia_semana = datos["fraud_por_dia_semana"]
    legit_por_dia_semana = datos["legit_por_dia_semana"]
    fraud_por_semana = datos["fraud_por_semana"]
    legit_por_semana = datos["legit_por_semana"]
    fraud_por_mes = datos["fraud_por_mes"]
    legit_por_mes = datos["legit_por_mes"]

    # Crear un DataFrame de pandas con los datos
    
    categorias = ["Total Registros", "Total Fraudulentas", "Porcentaje Fraudulentas", "Total Legitimas", "Porcentaje Legitimas", " "]
    valores = [total_registros, total_fraud, porcentaje_fraud, total_legit, porcentaje_legit, ' ']

    # Agregar los datos por hora del día
    for item in fraud_por_hora:
        categorias.append(f"Fraudes por Hora {item['hora']}")
        valores.append(item['count'])
            
    for item in legit_por_hora:
        categorias.append(f"Legitimas por Hora {item['hora']}")
        valores.append(item['count'])
    
    categorias.append(' ')
    valores.append(' ')

    # Agregar los datos por día de la semana
    for item in fraud_por_dia_semana:
        categorias.append(f"Fraudes por Día {item['dia_semana']}")
        valores.append(item['count'])
    for item in legit_por_dia_semana:
        categorias.append(f"Legitimas por Día {item['dia_semana']}")
        valores.append(item['count'])

    categorias.append(' ')
    valores.append(' ')
    # Agregar los datos por semana

    for item in fraud_por_semana:
        categorias.append(f"Fraudes por Semana {item['semana']}")
        valores.append(item['count'])
    for item in legit_por_semana:
        categorias.append(f"Legitimas por Semana {item['semana']}")
        valores.append(item['count'])

    categorias.append(' ')
    valores.append(' ')
    # Agregar los datos por mes
 
    for item in fraud_por_mes:
        categorias.append(f"Fraudes por Mes {item['mes']}")
        valores.append(item['count'])
    for item in legit_por_mes:
        categorias.append(f"Legitimas por Mes {item['mes']}")
        valores.append(item['count'])

    
    df = pd.DataFrame({ "Categoría": categorias, "Valores":  valores})
    
    # Crear un libro de trabajo de Excel
    wb = Workbook()
    # Seleccionar la hoja de cálculo activa (por defecto es la primera hoja)
    ws = wb.active

    # Definir los estilos de celda
    header_fill = PatternFill(start_color="1e4d70", end_color="1e4d70", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    value_fill = PatternFill(start_color="72e082", end_color="72e082", fill_type="solid")
    value_font = Font(color="FFFFFF")
    border = Border(left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"))

    # Aplicar formato a la primera fila (encabezados)
    for col_num, value in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Aplicar formato a las celdas de valores
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = value_fill
            # cell.font = value_font
            cell.border = border

    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Guardar el libro de trabajo
    archivo_excel = os.path.join('reporteria', 'reporte.xlsx')
    df.to_excel(archivo_excel, index=False)

    # Leer el archivo Excel generado
    with open(archivo_excel, 'rb') as file:
        excel_content = file.read()

    # Configurar la respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'
    response.write(excel_content)

    return response


@csrf_exempt
@api_view(['GET'])
def reporte_datos(request):
    required_fields = ['fechaInicial', 'fechaFinal', 'token', 'id']
    
    for field in required_fields:
        if field not in request.data:
            return JsonResponse({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
        
    
    # token = request.data['token']
    # try:
    #     empresa = Empresa.objects.get( token=token)
    # except Empresa.DoesNotExist:
    #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # plan_activo = revisarplan(empresa.id)

    # # Comprobar si el plan ha expirado
    # if not plan_activo:
    #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
    #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
    
    
    formato_esperado = '%Y-%m-%d'
    
    try:
        datetime.strptime(request.data['fechaInicial'], formato_esperado)
        datetime.strptime(request.data['fechaFinal'], formato_esperado)
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha no valido(El formato debe ser yyyy-mm-dd, ejemplo: 2023-05-11)'}, status=status.HTTP_401_UNAUTHORIZED)
    
    datos = reporte(1, request.data[ 'fechaInicial'], request.data['fechaFinal'])
    
    total_registros = datos["total_registros"]
    total_fraud = datos["total_fraud"]
    porcentaje_fraud = datos["porcentaje_fraud"]
    total_legit = datos["total_legit"]
    porcentaje_legit = datos["porcentaje_legit"]
    fraud_por_hora = datos["fraud_por_hora"]
    legit_por_hora = datos["legit_por_hora"]
    fraud_por_dia_semana = datos["fraud_por_dia_semana"]
    legit_por_dia_semana = datos["legit_por_dia_semana"]
    fraud_por_semana = datos["fraud_por_semana"]
    legit_por_semana = datos["legit_por_semana"]
    fraud_por_mes = datos["fraud_por_mes"]
    legit_por_mes = datos["legit_por_mes"]
    
    fraudesHora = []
    for item in fraud_por_hora:
        fraudesHora.append({f"Hora {item['hora']}": {item['count']}})

    legitimasHora = []
    for item in legit_por_hora:
        legitimasHora.append({f"Hora {item['hora']}": {item['count']}})
        
    fraudedia = []
    for item in fraud_por_dia_semana:
        fraudedia.append({f"Dia {item['dia_semana']}": {item['count']}})
        
    legitimasdia = []
    for item in legit_por_dia_semana:
        legitimasdia.append({f"Dia {item['dia_semana']}": {item['count']}})
    
    legitimassemana = []
    for item in legit_por_semana:
        legitimassemana.append({f"Semana {item['semana']}": {item['count']}})
    
    fraudesemana = []
    for item in fraud_por_semana:
        fraudesemana.append({f"Semana {item['semana']}": {item['count']}})
        
    legitimasmes = []
    for item in fraud_por_mes:
        legitimasmes.append({f"Mes {item['mes']}": {item['count']}})
        
    fraudesmes = []
    for item in legit_por_mes:
        fraudesmes.append({f"Mes {item['mes']}": {item['count']}})
        
        
    respuesta = {
        "Total Transacciones verificadas": total_registros,
        "Total transacciones fraudulentas":  total_fraud,
        "Porcentaje de transacciones fraudulentas":  porcentaje_fraud,
        "Total transacciones legitimas":  total_legit,
        "Porcentaje de transacciones legitimas":  porcentaje_legit,
        "Transacciones fraudulentas por hora": fraudesHora,
        "Transacciones legitimas por hora": legitimasHora,
        "Transacciones fraudulentas por dia de la semana": fraudedia,
        "Transacciones legitimas por dia de la semana": legitimasdia,
        "Transacciones fraudulentas por semana": fraudesemana,
        "Transacciones legitimas por semana": legitimassemana,
        "Transacciones fraudulentas por mes": legitimasmes,
        "Transacciones legitimas por mes": fraudesmes,
    }

    return Response(respuesta)


@csrf_exempt
def reporte_unico(request, token, id):
    required_fields = ['token', 'id']

    for field in required_fields:
        if field not in request.GET:
            return Response({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
    
    # token = request.data['token']
    # try:
    #     empresa = Empresa.objects.get( token=token)
    # except Empresa.DoesNotExist:
    #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # plan_activo = revisarplan(empresa.id)

    # # Comprobar si el plan ha expirado
    # if not plan_activo:
    #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
    #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
    transaccion = Verificaciones.objects.get(id=id)
    
    respuesta = {"Transaccion": transaccion.id, "Resultado": transaccion.resultado}
    
    return Response(respuesta)


@csrf_exempt
@api_view(['GET'])
def reporte_rango(request, token, inicio_id, fin_id):

    required_fields = ['token', 'inicio_id', 'fin_id']

    for field in required_fields:
        if field not in request.GET:
            return Response({'error': f'El campo {field} es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

    # try:
    #     empresa = Empresa.objects.get( token=token)
    # except Empresa.DoesNotExist:
    #     return JsonResponse({'error': 'Token de empresa inválido.'}, status=status.HTTP_401_UNAUTHORIZED)
    
    # plan_activo = revisarplan(empresa.id)

    # # Comprobar si el plan ha expirado
    # if not plan_activo:
    #     # Plan expirado, devolver una respuesta indicando que el plan ha expirado
    #         return JsonResponse({'error': f'El plan ha expirado. Por favor actualice su plan'}, status=status.HTTP_400_BAD_REQUEST)
    
    transacciones = Verificaciones.objects.filter(id__gt=inicio_id, id__lt=fin_id)
    
    respuesta = []
    for transaccion in transacciones:
        respuesta.append({"Transaccion": transaccion.id, "Resultado": transaccion.resultado})
        
    return Response(respuesta)